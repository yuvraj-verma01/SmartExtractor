from __future__ import annotations

from datetime import datetime
import json
import math
import shutil
from pathlib import Path
from typing import Any, Dict

from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

from app.core import audit, io_utils, jobs, pipeline, schema, state
from app.core.paths import EXPORTS_ROOT, job_paths


app = FastAPI(title="Lease extractor by Yuvraj Verma", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "time": datetime.utcnow().isoformat() + "Z"}


@app.post("/jobs")
async def create_job(
    request: Request,
    file: UploadFile | None = File(default=None),
    source_path: str | None = Form(default=None),
    name: str | None = Form(default=None),
) -> dict:
    paths = jobs.new_job()

    payload: Dict[str, Any] = {}
    if file is None and source_path is None:
        try:
            payload = await request.json()
        except Exception:
            payload = {}

    source_path = source_path or payload.get("source_path")
    if name is None:
        name = payload.get("name")

    if file is not None:
        paths.input_dir.mkdir(parents=True, exist_ok=True)
        dest = paths.input_pdf
        with dest.open("wb") as handle:
            handle.write(await file.read())
    elif source_path:
        src = Path(source_path)
        if not src.exists():
            raise HTTPException(status_code=400, detail="source_path not found")
        paths.input_dir.mkdir(parents=True, exist_ok=True)
        dest = paths.input_pdf
        dest.write_bytes(src.read_bytes())

    state.init_working_state(paths, llm_status="unknown")
    meta = jobs.load_job_meta(paths.root.name)
    if name:
        meta["name"] = name.strip()[:200]
        jobs.save_job_meta(paths.root.name, meta)
    return {"job": meta}


@app.get("/jobs")
def list_jobs() -> dict:
    return {"jobs": jobs.list_jobs()}


def _load_job_or_404(job_id: str):
    if not jobs.job_exists(job_id):
        raise HTTPException(status_code=404, detail="job not found")
    return job_paths(job_id)


@app.post("/jobs/{job_id}/run")
def run_job(job_id: str, background_tasks: BackgroundTasks) -> dict:
    paths = _load_job_or_404(job_id)
    meta = jobs.load_job_meta(paths.root.name)
    if meta.get("status") == "running":
        return {"job": meta, "message": "already_running"}
    meta["status"] = "running"
    meta.setdefault("pipeline", {})
    for stage in ["stage1", "stage2", "stage3"]:
        meta["pipeline"].setdefault(stage, {"status": "pending", "message": None})
    jobs.save_job_meta(paths.root.name, meta)
    background_tasks.add_task(pipeline.run_pipeline, paths.root.name)
    return {"job": meta, "message": "started"}


@app.post("/jobs/{job_id}/run_stage/{stage}")
async def run_job_stage(job_id: str, stage: str, request: Request, background_tasks: BackgroundTasks) -> dict:
    paths = _load_job_or_404(job_id)
    if stage not in {"stage1", "stage2", "stage3"}:
        raise HTTPException(status_code=400, detail="invalid stage")

    payload: Dict[str, Any] = {}
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    model = payload.get("model")

    meta = jobs.load_job_meta(paths.root.name)
    if meta.get("status") == "running":
        return {"job": meta, "message": "already_running"}

    meta["status"] = "running"
    meta.setdefault("pipeline", {})
    for s in ["stage1", "stage2", "stage3"]:
        meta["pipeline"].setdefault(s, {"status": "pending", "message": None})
    if stage == "stage3" and model:
        meta["llm_model"] = model
    jobs.save_job_meta(paths.root.name, meta)

    if stage == "stage1":
        background_tasks.add_task(pipeline.run_stage1, paths.root.name)
    elif stage == "stage2":
        background_tasks.add_task(pipeline.run_stage2, paths.root.name)
    else:
        background_tasks.add_task(pipeline.run_stage3, paths.root.name, model)
    return {"job": meta, "message": "started"}


@app.get("/jobs/{job_id}/state")
def job_state(job_id: str) -> dict:
    paths = _load_job_or_404(job_id)
    meta = jobs.load_job_meta(job_id)
    working_state = state.load_working_state(paths)
    if meta.get("llm_status"):
        working_state["llm_status"] = meta.get("llm_status")
    reviewed = sum(
        1 for f in working_state.get("fields", {}).values() if f.get("review", {}).get("status") == "reviewed"
    )
    total = len(schema.FIELDS)
    final_exists = (paths.final_dir / "lease_final.json").exists()
    evidence_by_field: Dict[str, list] = {field: [] for field in schema.FIELDS}

    def _dedupe(items: list) -> list:
        seen = set()
        out = []
        for it in items:
            if not isinstance(it, dict):
                continue
            key = (it.get("text") or it.get("snippet") or "", it.get("page"), it.get("line_no"), it.get("source_field"))
            if key in seen:
                continue
            seen.add(key)
            out.append(it)
        return out

    review_queue_path = paths.stage2_dir / "review_queue.json"
    review_queue = io_utils.read_json(review_queue_path, default={}) or {}
    items = review_queue.get("items", []) or []
    for it in items:
        field = it.get("field")
        if not field:
            continue
        evidence = it.get("evidence") or {}
        snippets = evidence.get("snippets") or []
        if snippets:
            evidence_by_field.setdefault(field, []).extend(snippets)

    anchors_path = paths.stage2_dir / "lease_anchors.json"
    anchors = io_utils.read_json(anchors_path, default={}) or {}
    for field in schema.FIELDS:
        if evidence_by_field.get(field):
            continue
        hits = anchors.get(field) or []
        if hits:
            for h in hits[:8]:
                if not isinstance(h, dict):
                    continue
                evidence_by_field.setdefault(field, []).append(
                    {
                        "text": h.get("snippet") or "",
                        "page": h.get("page"),
                        "line_no": h.get("line_no"),
                        "source_field": field,
                        "score": None,
                    }
                )

    extracted_path = paths.stage2_dir / "lease_extracted.json"
    extracted = io_utils.read_json(extracted_path, default={}) or {}
    extracted_evidence = extracted.get("evidence") or {}
    for field in schema.FIELDS:
        if evidence_by_field.get(field):
            continue
        ev = extracted_evidence.get(field) or {}
        text = ev.get("evidence")
        if text:
            evidence_by_field.setdefault(field, []).append(
                {
                    "text": text,
                    "page": ev.get("page"),
                    "line_no": ev.get("line_no"),
                    "source_field": "extracted",
                    "score": None,
                }
            )

    for field, items in list(evidence_by_field.items()):
        evidence_by_field[field] = _dedupe(items)
    return {
        "job": meta,
        "schema": schema.FIELDS,
        "working_state": working_state,
        "review_progress": {"reviewed": reviewed, "total": total},
        "final_exists": final_exists,
        "evidence_by_field": evidence_by_field,
    }


@app.post("/jobs/{job_id}/field_action")
def field_action(job_id: str, payload: Dict[str, Any]) -> dict:
    paths = _load_job_or_404(job_id)
    field = payload.get("field")
    action = payload.get("action")
    value = payload.get("value")
    source = payload.get("source")

    if field not in schema.FIELDS:
        raise HTTPException(status_code=400, detail="unknown field")
    if action not in {"accept", "edit", "clear"}:
        raise HTTPException(status_code=400, detail="unknown action")

    working_state = state.load_working_state(paths)
    fields = working_state.get("fields", {})
    entry = fields.get(field)
    if entry is None:
        raise HTTPException(status_code=400, detail="field missing in state")

    old_value = entry.get("value")
    if action == "clear":
        new_value = None
    elif value is not None:
        new_value = schema.coerce_value(field, value)
    else:
        new_value = old_value

    entry["value"] = new_value
    entry["review"] = {
        "status": "reviewed",
        "action": action,
        "reviewed_at": datetime.utcnow().isoformat() + "Z",
        "source": source,
    }

    state.save_working_state(paths, working_state)
    audit.append_action(
        paths,
        {
            "field": field,
            "action": action,
            "old_value": old_value,
            "new_value": new_value,
            "source": source,
        },
    )
    return {"field": field, "value": new_value, "review": entry["review"]}


def _cell_value(value: Any):
    if value is None:
        return None
    if isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    return json.dumps(value, ensure_ascii=True)


def _excel_export_path() -> Path:
    EXPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    return EXPORTS_ROOT / "lease_jobs.xlsx"


def _load_export_workbook() -> tuple:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"openpyxl missing: {exc}")

    export_path = _excel_export_path()
    if export_path.exists():
        wb = load_workbook(export_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lease Jobs"

    headers = ["job_id", "job_name", "finalized_at"] + schema.FIELDS
    existing = [ws.cell(row=1, column=idx + 1).value for idx in range(len(headers))] if ws.max_row >= 1 else []
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
    return wb, ws, headers, export_path


def _find_job_row(ws, job_id: str) -> int | None:
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == job_id:
            return row_idx
    return None


def _export_job(job_id: str, require_review: bool = True) -> dict:
    paths = _load_job_or_404(job_id)
    working_state = state.load_working_state(paths)
    fields = working_state.get("fields", {})

    if require_review:
        unreviewed = [name for name, f in fields.items() if f.get("review", {}).get("status") != "reviewed"]
        if unreviewed:
            raise HTTPException(status_code=400, detail={"unreviewed": unreviewed})

    row = {name: f.get("value") for name, f in fields.items()}
    final = {
        "row": row,
        "audit_log": audit.read_audit_log(paths),
        "source": {
            "lease_validated": str((paths.stage2_dir / "lease_validated.json").resolve())
            if (paths.stage2_dir / "lease_validated.json").exists()
            else None,
            "review_queue": str((paths.stage2_dir / "review_queue.json").resolve())
            if (paths.stage2_dir / "review_queue.json").exists()
            else None,
            "llm_suggestions": str((paths.stage3_dir / "lease_llm_suggestions.json").resolve())
            if (paths.stage3_dir / "lease_llm_suggestions.json").exists()
            else None,
        },
    }

    io_utils.write_json(paths.final_dir / "lease_final.json", final)

    meta = jobs.load_job_meta(job_id)
    exported_at = datetime.utcnow().isoformat() + "Z"
    wb, ws, headers, export_path = _load_export_workbook()
    row_idx = _find_job_row(ws, job_id)
    if row_idx is None:
        row_idx = ws.max_row + 1
        if row_idx < 2:
            row_idx = 2
    values = [job_id, meta.get("name"), exported_at] + [row.get(field) for field in schema.FIELDS]
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=row_idx, column=col_idx).value = _cell_value(value)
    wb.save(export_path)

    meta["excel_row"] = row_idx
    meta["excel_exported_at"] = exported_at
    if require_review:
        meta["status"] = "finalized"
        meta["finalized_at"] = exported_at
    jobs.save_job_meta(job_id, meta)
    return meta


@app.post("/jobs/{job_id}/save")
def save_changes(job_id: str, payload: Dict[str, Any]) -> dict:
    paths = _load_job_or_404(job_id)
    fields_payload = payload.get("fields") if isinstance(payload, dict) else None
    if not isinstance(fields_payload, dict):
        raise HTTPException(status_code=400, detail="fields payload required")

    working_state = state.load_working_state(paths)
    fields = working_state.get("fields", {})
    updated_fields = []
    for field, raw in fields_payload.items():
        if field not in schema.FIELDS:
            continue
        entry = fields.get(field)
        if entry is None:
            continue
        entry["value"] = schema.coerce_value(field, raw)
        updated_fields.append(field)

    working_state["fields"] = fields
    state.save_working_state(paths, working_state)
    meta = jobs.load_job_meta(job_id)
    meta["saved_at"] = working_state.get("updated_at")
    jobs.save_job_meta(job_id, meta)
    return {"job": meta, "updated_fields": updated_fields, "saved_at": working_state.get("updated_at")}


@app.post("/jobs/{job_id}/export_excel")
def export_excel(job_id: str) -> dict:
    meta = _export_job(job_id, require_review=True)
    return {"job": meta}


@app.post("/jobs/{job_id}/finalize")
def finalize_job(job_id: str) -> dict:
    meta = _export_job(job_id, require_review=True)
    return {"job": meta}


@app.get("/jobs/{job_id}/download/final_json")
def download_final_json(job_id: str):
    paths = _load_job_or_404(job_id)
    final_path = paths.final_dir / "lease_final.json"
    if not final_path.exists():
        raise HTTPException(status_code=404, detail="final_json not found")
    return FileResponse(final_path, filename="lease_final.json")


@app.get("/jobs/{job_id}/download/working_json")
def download_working_json(job_id: str):
    paths = _load_job_or_404(job_id)
    working_state_path = paths.working_state_path
    if not working_state_path.exists():
        raise HTTPException(status_code=404, detail="working_state not found")
    return FileResponse(working_state_path, filename="working_state.json")


@app.delete("/jobs/{job_id}")
def delete_job(job_id: str) -> dict:
    paths = _load_job_or_404(job_id)
    meta = jobs.load_job_meta(job_id)
    if meta.get("status") == "running":
        raise HTTPException(status_code=400, detail="job is running")
    shutil.rmtree(paths.root, ignore_errors=True)
    return {"deleted": job_id}


@app.get("/jobs/{job_id}/download/xlsx")
def download_xlsx(job_id: str):
    _load_job_or_404(job_id)
    export_path = _excel_export_path()
    if not export_path.exists():
        raise HTTPException(status_code=404, detail="no excel export yet")
    return FileResponse(export_path, filename="lease_jobs.xlsx")
