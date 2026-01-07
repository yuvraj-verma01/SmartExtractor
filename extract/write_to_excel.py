# extract/write_to_excel.py
"""
Stage 3B: Append the final row to Excel (local)

Reads:
  data/outputs/lease_final.json

Writes/Appends:
  data/outputs/leases.xlsx   (default)
    - creates workbook if missing
    - creates sheet if missing
    - writes headers if empty
    - appends one row

Run:
  python .\\extract\\write_to_excel.py
Optional:
  python .\\extract\\write_to_excel.py --xlsx "C:\\path\\to\\your.xlsx" --sheet "Sheet1"
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook


OUT_DIR = Path("data/outputs")
FINAL_PATH = OUT_DIR / "lease_final.json"

DEFAULT_XLSX = OUT_DIR / "leases.xlsx"
DEFAULT_SHEET = "Leases"

# Your column order (from your message)
COLUMNS = [
    "city",
    "building_name",
    "floors_units",
    "lease_start_date",
    "lease_end_date",
    "rent_start_date",
    "handover_date",
    "lease_tenure_months",
    "lock_in_period",
    "lock_in_end_date",
    "rent_free_period_months",
    "termination_notice_period_months",
    "renewal_notice_period_months",
    "carpet_area_sqft",
    "super_builtup_area_sqft",
    "efficiency",
    "cam_area_sqft",
    "parking_4w_included",
    "parking_2w_included",
    "monthly_cam_rs",
    "monthly_rent_rs",
    "rate_per_sqft_rs",
    "parking_charges_rs",
    "renewal_option",
    "stamp_duty_rs",
    "ifrsd_rs",
]


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    # remove default sheet if we’ll create our own
    return wb


def ensure_sheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    # If workbook is new and has default "Sheet", reuse it by renaming.
    if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        ws = wb["Sheet"]
        ws.title = sheet_name
        return ws
    return wb.create_sheet(sheet_name)


def ensure_headers(ws, headers: List[str]) -> None:
    # If first row is empty, write headers
    if ws.max_row == 1 and all(ws.cell(row=1, column=i + 1).value is None for i in range(len(headers))):
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = h
        return

    # If sheet is totally empty (rare)
    if ws.max_row == 0:
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i).value = h


def append_row(ws, headers: List[str], row: Dict[str, Any]) -> int:
    next_row = ws.max_row + 1
    for i, h in enumerate(headers, 1):
        ws.cell(row=next_row, column=i).value = row.get(h, None)
    return next_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=str, default=str(DEFAULT_XLSX), help="Path to .xlsx file")
    ap.add_argument("--sheet", type=str, default=DEFAULT_SHEET, help="Sheet name")
    args = ap.parse_args()

    if not FINAL_PATH.exists():
        raise FileNotFoundError(f"Missing {FINAL_PATH}. Run review_loop.py first.")

    final = read_json(FINAL_PATH)
    row: Dict[str, Any] = final.get("row", {}) or {}

    xlsx_path = Path(args.xlsx)
    wb = ensure_workbook(xlsx_path)
    ws = ensure_sheet(wb, args.sheet)
    ensure_headers(ws, COLUMNS)
    r = append_row(ws, COLUMNS, row)

    # Make sure at least one sheet is visible
    wb.active = ws

    # Save
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)

    print(f" Appended to Excel: {xlsx_path} (sheet={args.sheet}, row={r})")


if __name__ == "__main__":
    main()
